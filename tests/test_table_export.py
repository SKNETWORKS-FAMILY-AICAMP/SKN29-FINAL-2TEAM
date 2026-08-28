"""`table_export` 가 굽는 xlsx 의 계약.

여기서 보는 것은 **모델이 준 값을 그대로 믿지 않는다**는 한 가지다 —
수식으로 해석될 글자, 열 수가 안 맞는 행, 엑셀이 거절하는 시트 이름.
"""

from io import BytesIO
from datetime import datetime
from unittest.mock import patch

from django.test import SimpleTestCase
from openpyxl import load_workbook

from services.document_export import build_xlsx
from services.harness.registry import _XLSX_MIME


def _sheet(data: bytes):
    return load_workbook(BytesIO(data)).active


class BuildXlsxTests(SimpleTestCase):
    def test_머리글과_행을_그대로_적는다(self):
        sheet = _sheet(
            build_xlsx(
                title="업무 목록",
                columns=["제목", "담당", "공수"],
                rows=[["로그인 구현", "김준억", 16], ["정산 배치", "원빈", 8]],
            )
        )

        self.assertEqual(sheet.title, "업무 목록")
        self.assertEqual([c.value for c in sheet[1]], ["제목", "담당", "공수"])
        self.assertEqual([c.value for c in sheet[2]], ["로그인 구현", "김준억", 16])
        # 숫자는 숫자로 남아야 한다 — 문자열이면 합계도 정렬도 안 된다.
        self.assertIsInstance(sheet.cell(row=2, column=3).value, int)

    def test_수식으로_보이는_글자를_실행하지_않는다(self):
        """문서에 적혀 있던 `=...` 이 셀 수식이 되면 파일을 여는 사람 쪽에서 돈다."""

        sheet = _sheet(
            build_xlsx(
                title="근거",
                columns=["=HYPERLINK(\"http://x\")"],
                rows=[["=1+1"], ["+1"], ["@SUM(A1)"], ["-2"]],
            )
        )

        for row in range(1, 5 + 1):
            cell = sheet.cell(row=row, column=1)
            self.assertEqual(cell.data_type, "s", f"{row}행이 글자가 아니다: {cell.value!r}")
        # 값 자체는 안 바꾼다. 보이는 글자가 원문과 달라지면 그것도 왜곡이다.
        self.assertEqual(sheet.cell(row=2, column=1).value, "=1+1")

    def test_http_url은_클릭_가능한_하이퍼링크로_남긴다(self):
        sheet = _sheet(
            build_xlsx(
                title="출처",
                columns=["이름", "URL"],
                rows=[
                    ["공식 문서", "https://example.com/docs?q=1"],
                    ["안전하지 않은 스킴", "javascript:alert(1)"],
                    ["문장 속 주소", "참고: https://example.com"],
                ],
            )
        )

        self.assertEqual(sheet["B2"].value, "https://example.com/docs?q=1")
        self.assertEqual(sheet["B2"].hyperlink.target, "https://example.com/docs?q=1")
        self.assertEqual(sheet["B2"].font.underline, "single")
        self.assertIsNone(sheet["B3"].hyperlink)
        self.assertIsNone(sheet["B4"].hyperlink)

    def test_출처명과_http_url을_한_셀의_하이퍼링크로_남긴다(self):
        sheet = _sheet(
            build_xlsx(
                title="출처",
                columns=["출처"],
                rows=[["[공식 홈페이지](https://example.com/company)"]],
            )
        )

        self.assertEqual(sheet["A2"].value, "공식 홈페이지")
        self.assertEqual(sheet["A2"].hyperlink.target, "https://example.com/company")

    def test_열_수가_안_맞는_행을_머리글에_맞춘다(self):
        sheet = _sheet(
            build_xlsx(
                title="t",
                columns=["A", "B", "C"],
                rows=[["1"], ["1", "2", "3", "4"]],
            )
        )

        self.assertEqual([c.value for c in sheet[2]], ["1", None, None])
        self.assertEqual([c.value for c in sheet[3]], ["1", "2", "3"])

    def test_행이_없어도_머리글만_있는_파일이_나온다(self):
        sheet = _sheet(build_xlsx(title="t", columns=["A", "B"], rows=[]))

        self.assertEqual(sheet.max_row, 1)
        self.assertEqual([c.value for c in sheet[1]], ["A", "B"])

    def test_엑셀이_거절하는_시트_이름을_고친다(self):
        # 금지 글자(`[]:*?/\`)와 31자 상한. 어기면 파일이 아예 안 열린다.
        long_title = "가" * 40
        self.assertEqual(len(_sheet(build_xlsx(title=long_title, columns=["A"], rows=[])).title), 31)
        self.assertEqual(_sheet(build_xlsx(title="a/b:c", columns=["A"], rows=[])).title, "abc")
        # 금지 글자만 있던 제목은 빈 이름이 되는데, 엑셀은 그것도 거절한다.
        self.assertEqual(_sheet(build_xlsx(title="///", columns=["A"], rows=[])).title, "표")

    def test_참거짓과_빈값을_사람이_읽는_모양으로_적는다(self):
        sheet = _sheet(
            build_xlsx(title="t", columns=["A", "B", "C"], rows=[[True, False, None]])
        )

        # bool 은 int 의 하위형이라 순서를 안 지키면 1/0 으로 들어간다.
        self.assertEqual([c.value for c in sheet[2]], ["예", "아니오", None])

    def test_큰_표를_읽기_위한_기본_서식을_적용한다(self):
        sheet = _sheet(
            build_xlsx(
                title="진행 현황",
                columns=["상태", "시작일", "진행률", "비고"],
                rows=[["진행 중", "2026-08-28", 0.85, "긴 설명"]],
            )
        )

        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertIsNone(sheet.auto_filter.ref)
        self.assertEqual(list(sheet.tables), ["ExportedData"])
        self.assertEqual(sheet.tables["ExportedData"].autoFilter.ref, "A1:D2")
        self.assertEqual(sheet["A1"].fill.fgColor.rgb, "00243B67")
        self.assertEqual(sheet["A1"].font.color.rgb, "00FFFFFF")
        self.assertTrue(sheet.sheet_view.showGridLines)

        self.assertIsInstance(sheet["B2"].value, datetime)
        self.assertEqual(sheet["B2"].number_format, "yyyy-mm-dd")
        self.assertEqual(sheet["C2"].number_format, "0%")
        self.assertTrue(sheet["D2"].alignment.wrap_text)


class TableExportToolTests(SimpleTestCase):
    """도구 핸들러가 「내 파일」에 무엇을 남기는가."""

    def _call(self, **kwargs):
        from services.harness.registry import _table_export

        defaults = {
            "account_id": "AC001",
            "title": "업무 목록",
            "columns": ["제목", "공수"],
            "rows": [["로그인 구현", 16]],
        }
        return _table_export(**{**defaults, **kwargs})

    @patch("services.harness.registry.DocumentRepository.mark_stored")
    @patch("services.harness.registry.storage.save", return_value="sha256:abc123def4567890")
    @patch(
        "services.harness.registry.PersonalDocumentRepository.create_generated",
        return_value="DC099",
    )
    def test_생성_문서로_저장한다(self, create, save, mark_stored):
        result = self._call()

        # 업로드가 아니라 생성이다 — `create` 를 부르면 목록에서 안 갈린다.
        create.assert_called_once()
        self.assertEqual(create.call_args.kwargs["mime_type"], _XLSX_MIME)
        self.assertTrue(create.call_args.kwargs["file_name"].endswith(".xlsx"))

        # 개인 소유 자리에 둔다. 팀 키에 두면 팀을 지울 때 함께 지워진다.
        key = save.call_args.args[0]
        self.assertTrue(key.startswith("user/AC001/DC099"), key)
        self.assertTrue(key.endswith(".xlsx"), key)

        # 저장된 바이트가 실제로 열리는 xlsx 여야 한다.
        sheet = _sheet(save.call_args.args[1])
        self.assertEqual([c.value for c in sheet[2]], ["로그인 구현", 16])

        # **`file` 키가 채팅 다운로드 카드의 계약이다**(2026-08-26). 평평한
        # `doc_id` 로 두면 `document_search` 결과의 `doc_id` 와 구별되지 않는다.
        self.assertEqual(
            result["file"], {"doc_id": "DC099", "file_name": create.call_args.kwargs["file_name"], "mime_type": _XLSX_MIME}
        )
        self.assertEqual(result["rows"], 1)
        # 표 내용을 되돌려주지 않는다 — 모델이 방금 보낸 값이다.
        self.assertNotIn("data", result)

    @patch("services.harness.registry.DocumentRepository.mark_stored")
    @patch("services.harness.registry.storage.save", return_value="sha256:abc")
    @patch(
        "services.harness.registry.PersonalDocumentRepository.create_generated",
        return_value="DC099",
    )
    def test_파일_이름에_경로_구분자를_넣지_않는다(self, create, save, mark_stored):
        """제목이 곧 파일 이름이 된다. `/` 가 남으면 저장 경로가 갈라진다."""

        self._call(title="8/26 주간 보고")
        self.assertNotIn("/", create.call_args.kwargs["file_name"])

    def test_사람이_고칠_수_있는_실패는_그대로_말한다(self):
        from services.harness.registry import ToolInputError

        with self.assertRaises(ToolInputError):
            self._call(columns=[])
        # 모델이 행을 dict 로 보내는 실수. 조용히 빈 칸으로 두면 안 된다.
        with self.assertRaises(ToolInputError):
            self._call(rows=[{"제목": "x"}])
