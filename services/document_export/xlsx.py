"""표 하나를 xlsx 바이트로 굽는다.

`table_export` 도구가 부른다. 표의 출처는 **모델이 넘긴 인자**이고, 그 값은
사용자 문서에서 온 글자다 — 그래서 아래 두 가지를 신뢰하지 않는다.

1. **수식으로 해석될 수 있는 글자.** openpyxl 은 `=` 로 시작하는 문자열을
   수식(`data_type='f'`)으로 추론한다(2026-08-26 실측). 문서에 적혀 있던
   `=SUM(...)` 이 그대로 살아 있는 셀이 되면, 파일을 여는 사람의 엑셀에서
   실행된다. **모든 글자 셀의 타입을 `'s'` 로 못박아** 값이 곧 글자가 되게 한다.
2. **행마다 다른 열 구성.** 모델이 만든 배열이라 어떤 행은 열이 모자라고 어떤
   행은 남는다. `columns` 를 기준으로 자르고 채운다 — 여기서 맞추지 않으면
   엑셀에서 열이 밀린 채 열린다.
"""

from __future__ import annotations

import re
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

#: 시트 이름 상한(엑셀 규격). 넘으면 파일이 열리지 않는다.
_SHEET_TITLE_MAX = 31

#: 엑셀이 시트 이름에 못 쓰는 글자(규격). 제목이 그대로 시트 이름이 되므로 턴다.
_SHEET_TITLE_BANNED = set(r"[]:*?/\\")

#: 열 너비 상한. 긴 셀 하나가 화면을 다 먹는 것을 막는다.
_WIDTH_MAX = 42
_WIDTH_MIN = 8

_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_HTTP_URL = re.compile(r"^https?://[^\s]+$", re.IGNORECASE)
_MARKDOWN_LINK = re.compile(r"^\[([^\]\r\n]+)\]\((https?://[^\s)]+)\)$", re.IGNORECASE)
_DATE_HEADERS = ("날짜", "일자", "시작일", "마감일", "기한")
_PERCENT_HEADERS = ("진행률", "부하율", "달성률", "비율")

_HEADER_FILL = PatternFill("solid", fgColor="243B67")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_STATUS_FILLS = {
    "완료": PatternFill("solid", fgColor="E5F4EA"),
    "진행 중": PatternFill("solid", fgColor="E8F0FE"),
    "검수 중": PatternFill("solid", fgColor="FFF4D6"),
    "할 일": PatternFill("solid", fgColor="F1F3F5"),
}
_PRIORITY_FILLS = {
    "최우선": PatternFill("solid", fgColor="FDE8E7"),
    "높음": PatternFill("solid", fgColor="FFF0E0"),
}


def _sheet_title(title: str) -> str:
    cleaned = "".join(ch for ch in title if ch not in _SHEET_TITLE_BANNED).strip()
    # 빈 이름은 엑셀이 거절한다. 제목이 금지 글자뿐이었던 경우다.
    return cleaned[:_SHEET_TITLE_MAX] or "표"


def _cell_value(value: Any) -> tuple[Any, bool]:
    """셀에 넣을 값과 「글자로 못박아야 하는가」를 돌려준다.

    숫자는 숫자로 둔다 — 문자열로 넣으면 합계도 정렬도 안 된다. `bool` 을 먼저
    거르는 이유는 파이썬에서 `bool` 이 `int` 의 하위형이라, 나중에 보면 True 가
    1 로 들어가기 때문이다.
    """

    if value is None:
        return None, False
    if isinstance(value, bool):
        return ("예" if value else "아니오"), True
    if isinstance(value, (int, float)):
        return value, False
    return str(value), True


def _typed_value(header: str, value: Any) -> tuple[Any, bool, str | None]:
    """열 이름이 명확할 때만 날짜·백분율 표시 형식을 덧붙인다.

    모델 값의 의미를 추측해 바꾸지는 않는다. ISO 날짜 문자열만 실제 날짜로
    바꾸고, 0~1 범위 숫자만 백분율로 표시한다.
    """

    normalized = str(header).replace(" ", "")
    if isinstance(value, str) and _ISO_DATE.fullmatch(value) and any(
        token in normalized for token in _DATE_HEADERS
    ):
        return date.fromisoformat(value), False, "yyyy-mm-dd"
    if isinstance(value, (int, float)) and not isinstance(value, bool) and any(
        token in normalized for token in _PERCENT_HEADERS
    ) and 0 <= value <= 1:
        return value, False, "0%"
    cell_value, force_text = _cell_value(value)
    return cell_value, force_text, None


def build_xlsx(*, title: str, columns: list[str], rows: list[list[Any]]) -> bytes:
    """제목·머리글·행으로 xlsx 를 만들어 바이트로 돌려준다."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title(title)

    sheet.append(list(columns))
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # 머리글도 모델이 준 글자다. 본문과 같은 이유로 못박는다.
        if cell.value is not None:
            cell.data_type = "s"

    for row in rows:
        # 열 수를 머리글에 맞춘다 — 모자라면 채우고 남으면 자른다.
        padded = list(row[: len(columns)]) + [None] * max(0, len(columns) - len(row))
        written = sheet.max_row + 1
        for index, raw in enumerate(padded, start=1):
            value, force_text, number_format = _typed_value(columns[index - 1], raw)
            cell = sheet.cell(row=written, column=index, value=value)
            if force_text:
                # 여기가 수식 실행을 막는 한 줄이다. 위 모듈 주석 1번.
                cell.data_type = "s"
            # 셀 전체가 HTTP(S) URL일 때만 클릭 가능한 링크로 만든다. 임의의
            # 스킴이나 문장 속 URL은 연결하지 않아, 모델 입력이 엑셀의 외부
            # 동작으로 과도하게 확장되지 않게 한다. 표시 문자열은 원문 그대로다.
            markdown_link = _MARKDOWN_LINK.fullmatch(value.strip()) if isinstance(value, str) else None
            if markdown_link:
                cell.value = markdown_link.group(1)
                cell.hyperlink = markdown_link.group(2)
                cell.font = Font(color="0563C1", underline="single")
            elif isinstance(value, str) and _HTTP_URL.fullmatch(value.strip()):
                cell.hyperlink = value.strip()
                cell.font = Font(color="0563C1", underline="single")
            if number_format:
                cell.number_format = number_format
            cell.alignment = Alignment(
                horizontal="right" if isinstance(value, (int, float)) else "left",
                vertical="top",
                wrap_text=True,
            )
            text = str(raw).strip() if raw is not None else ""
            if text in _STATUS_FILLS:
                cell.fill = _STATUS_FILLS[text]
            elif text in _PRIORITY_FILLS:
                cell.fill = _PRIORITY_FILLS[text]

    # 폭은 그 열에서 제일 긴 글자에 맞춘다. 안 맞추면 전부 `#####` 로 보인다.
    for index in range(1, len(columns) + 1):
        longest = max(
            (len(str(cell.value)) for cell in sheet[get_column_letter(index)] if cell.value is not None),
            default=0,
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(
            _WIDTH_MAX, max(_WIDTH_MIN, longest + 2)
        )

    # 머리글을 고정한다. 행이 많으면 스크롤했을 때 어느 열인지 알 수 없다.
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 26
    # 표 바깥의 빈 셀까지 선이 사라지면 작은 표가 시트에서 고립되거나 깨진
    # 것처럼 보인다. Excel의 익숙한 탐색 기준을 유지하도록 기본 눈금선을 둔다.
    sheet.sheet_view.showGridLines = True

    # 필터와 줄무늬 행은 큰 표를 빠르게 훑기 위한 최소 장치다. 값이나 계산을
    # 추가하지 않고 같은 범위의 표현만 보강한다.
    end_column = get_column_letter(max(len(columns), 1))
    end_row = max(sheet.max_row, 1)
    table = Table(displayName="ExportedData", ref=f"A1:{end_column}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    # Excel Table 자체가 같은 범위의 AutoFilter를 가진다. 워크시트에도 별도
    # AutoFilter를 기록하면 Excel 데스크톱이 중복 정의로 판단해 표를 복구·제거한다.
    sheet.add_table(table)
    sheet.print_title_rows = "1:1"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.outlinePr.summaryBelow = True

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
