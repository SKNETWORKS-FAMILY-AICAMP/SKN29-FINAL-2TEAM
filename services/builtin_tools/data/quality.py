"""Frictionless·jsonschema·DuckDB 기반 데이터 품질 검사."""

from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

import duckdb
from frictionless import Resource, Schema, validate
from jsonschema import Draft202012Validator
from openpyxl import load_workbook

from services.builtin_tools.common.errors import BuiltinToolError
from services.builtin_tools.common.limits import MAX_FILE_BYTES
from services.builtin_tools.data.transformer import (
    CSV_MIME,
    JSON_MIME,
    PARQUET_MIME,
    SUPPORTED_TABLE_MIME_TYPES,
    _columns,
    _create_source,
    _identifier,
    _source_path,
)
from services.builtin_tools.documents.reader import XLSX_MIME


def _raw_rows(data: bytes, mime_type: str) -> tuple[list[str], list[list[Any]]]:
    if mime_type == CSV_MIME:
        try:
            rows = list(csv.reader(StringIO(data.decode("utf-8-sig"))))
        except UnicodeDecodeError as exc:
            raise BuiltinToolError("INVALID_CSV", "CSV는 UTF-8 인코딩이어야 합니다.") from exc
        return (rows[0], rows[1:]) if rows else ([], [])
    if mime_type == JSON_MIME:
        try:
            payload = json.loads(data)
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise BuiltinToolError("INVALID_JSON", "JSON 내용을 읽지 못했습니다.") from exc
        if not isinstance(payload, list) or any(not isinstance(item, dict) for item in payload):
            raise BuiltinToolError("INVALID_JSON_TABLE", "표 JSON은 객체 배열이어야 합니다.")
        columns = list(dict.fromkeys(key for item in payload for key in item))
        return columns, [[item.get(column) for column in columns] for item in payload]
    if mime_type == XLSX_MIME:
        try:
            workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
        except Exception as exc:  # noqa: BLE001 - parser 내부 경로를 숨긴다.
            raise BuiltinToolError("INVALID_XLSX", "XLSX 내용을 읽지 못했습니다.") from exc
        try:
            if len(workbook.sheetnames) != 1:
                raise BuiltinToolError(
                    "ONE_SHEET_REQUIRED", "데이터 품질 검사는 시트가 하나인 XLSX부터 지원합니다."
                )
            rows = [list(row) for row in workbook.active.iter_rows(values_only=True)]
            if not rows:
                return [], []
            return [str(value or "") for value in rows[0]], rows[1:]
        finally:
            workbook.close()
    if mime_type == PARQUET_MIME:
        with TemporaryDirectory(prefix="quality-parquet-") as raw_directory:
            path = Path(raw_directory) / "source.parquet"
            path.write_bytes(data)
            connection = duckdb.connect(":memory:")
            try:
                cursor = connection.execute("SELECT * FROM read_parquet(?)", [str(path)])
                columns = [item[0] for item in cursor.description]
                return columns, [list(row) for row in cursor.fetchall()]
            except duckdb.Error as exc:
                raise BuiltinToolError("INVALID_PARQUET", "Parquet 내용을 읽지 못했습니다.") from exc
            finally:
                connection.close()
    raise BuiltinToolError("UNSUPPORTED_FORMAT", "CSV, JSON, XLSX, Parquet만 지원합니다.")


def _frictionless_errors(
    columns: list[str], rows: list[list[Any]], schema_descriptor: dict[str, Any] | None
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not columns:
        return [{"code": "EMPTY_TABLE", "message": "열이 없는 빈 표입니다."}], {"fields": []}
    if len(columns) != len(set(columns)) or any(not name for name in columns):
        return [
            {"code": "INVALID_HEADER", "message": "열 이름이 비어 있거나 중복되었습니다."}
        ], {"fields": []}
    try:
        schema = Schema.from_descriptor(schema_descriptor) if schema_descriptor else None
        resource = Resource(data=[columns, *rows], schema=schema)
        if schema is None:
            resource.infer()
            inferred = resource.schema.to_descriptor()
            return [], inferred
        report = validate(resource)
    except Exception as exc:  # noqa: BLE001 - 라이브러리 내부 값을 숨긴다.
        raise BuiltinToolError("INVALID_SCHEMA", "데이터 검사 기준을 해석하지 못했습니다.") from exc

    errors = []
    for task in report.tasks:
        for error in task.errors:
            errors.append(
                {
                    "code": error.type,
                    "row": error.row_number,
                    "column": error.field_name,
                    "value": getattr(error, "cell", None),
                    "message": error.message,
                }
            )
    return errors, schema.to_descriptor()


def _json_schema_errors(data: bytes, descriptor: dict[str, Any] | None) -> list[dict[str, Any]]:
    if descriptor is None:
        return []
    try:
        payload = json.loads(data)
        validator = Draft202012Validator(descriptor)
    except Exception as exc:  # noqa: BLE001 - 스키마 라이브러리 내부 값을 숨긴다.
        raise BuiltinToolError("INVALID_JSON_SCHEMA", "JSON 또는 검사 기준을 해석하지 못했습니다.") from exc
    return [
        {
            "code": "json-schema",
            "location": list(error.absolute_path),
            "message": error.message,
        }
        for error in sorted(
            validator.iter_errors(payload), key=lambda item: str(list(item.absolute_path))
        )
    ]


def check_data_quality(
    *,
    data: bytes,
    mime_type: str,
    schema: dict[str, Any] | None = None,
    json_schema: dict[str, Any] | None = None,
    check_missing: bool = True,
    check_duplicates: bool = True,
    duplicate_keys: list[str] | None = None,
) -> dict[str, Any]:
    """원본을 바꾸지 않고 위치가 포함된 품질 오류를 반환한다."""

    if mime_type not in SUPPORTED_TABLE_MIME_TYPES:
        raise BuiltinToolError("UNSUPPORTED_FORMAT", "CSV, JSON, XLSX, Parquet만 지원합니다.")
    if not data:
        raise BuiltinToolError("EMPTY_FILE", "빈 표 파일은 검사할 수 없습니다.")
    if len(data) > MAX_FILE_BYTES:
        raise BuiltinToolError("FILE_TOO_LARGE", "표 파일은 20MB 이하여야 합니다.")

    # 표가 아닌 JSON(단일 객체 등)이라도 json_schema 만으로 검사할 수 있어야 한다
    # — 설정 파일 한 건을 스키마에 대조하는 것은 자연스러운 요청이다. 표 형태가
    # 아닌데 json_schema 도 없을 때만 "표가 아니다"로 거절한다.
    if mime_type == JSON_MIME:
        try:
            columns, rows = _raw_rows(data, mime_type)
        except BuiltinToolError as exc:
            if exc.code != "INVALID_JSON_TABLE" or json_schema is None:
                raise
            schema_errors = _json_schema_errors(data, json_schema)
            return {
                "valid": not schema_errors,
                "error_count": len(schema_errors),
                "errors": schema_errors,
                "inferred_schema": None,
                "note": "표 형태가 아니어서 JSON Schema 검사만 수행했습니다.",
            }
    else:
        columns, rows = _raw_rows(data, mime_type)
    errors, inferred_schema = _frictionless_errors(columns, rows, schema)
    if mime_type == JSON_MIME:
        errors.extend(_json_schema_errors(data, json_schema))

    # 누락과 정확한 중복은 DuckDB로 위치를 계산한다. 모델이 SQL을 전달하는 경로는 없다.
    with TemporaryDirectory(prefix="quality-table-") as raw_directory:
        directory = Path(raw_directory)
        path, source_mime = _source_path(
            data=data, mime_type=mime_type, directory=directory, name="source"
        )
        connection = duckdb.connect(":memory:")
        try:
            _create_source(connection, table="source", path=path, mime=source_mime)
            available = set(_columns(connection, "source"))
            if check_missing:
                for column in columns:
                    quoted = _identifier(column, available)
                    missing = connection.execute(
                        f'SELECT source_row FROM (SELECT row_number() OVER () + 1 AS source_row, '
                        f'* FROM "source") AS numbered WHERE {quoted} IS NULL '
                        f"OR trim(CAST({quoted} AS VARCHAR)) = ''"
                    ).fetchall()
                    errors.extend(
                        {
                            "code": "missing-value",
                            "row": int(item[0]),
                            "column": column,
                            "message": "값이 비어 있습니다.",
                        }
                        for item in missing
                    )
            if check_duplicates and rows:
                keys = duplicate_keys or columns
                key_sql = [_identifier(name, available) for name in keys]
                duplicate_rows = connection.execute(
                    f'SELECT {", ".join(key_sql)}, count(*) AS duplicate_count FROM "source" '
                    f'GROUP BY {", ".join(key_sql)} HAVING count(*) > 1'
                ).fetchall()
                for item in duplicate_rows:
                    errors.append(
                        {
                            "code": "duplicate-row",
                            "columns": keys,
                            "values": list(item[:-1]),
                            "count": int(item[-1]),
                            "message": "같은 값의 행이 중복되었습니다.",
                        }
                    )
        except duckdb.Error as exc:
            raise BuiltinToolError("QUALITY_CHECK_FAILED", "데이터 품질 검사를 완료하지 못했습니다.") from exc
        finally:
            connection.close()

    return {
        "valid": not errors,
        "error_count": len(errors),
        "errors": errors,
        "inferred_schema": inferred_schema,
    }
