"""DuckDB와 제한된 Query DSL을 사용하는 표 가공·집계."""

from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

import duckdb
from openpyxl import Workbook, load_workbook

from services.builtin_tools.common.errors import BuiltinToolError
from services.builtin_tools.common.json_values import json_value
from services.builtin_tools.common.limits import (
    MAX_FILE_BYTES,
    MAX_OUTPUT_BYTES,
    MAX_TABLE_EXPORT_ROWS,
)
from services.builtin_tools.documents.reader import XLSX_MIME
from services.builtin_tools.documents.package_safety import validate_ooxml_package

CSV_MIME = "text/csv"
JSON_MIME = "application/json"
PARQUET_MIME = "application/vnd.apache.parquet"
SUPPORTED_TABLE_MIME_TYPES = frozenset({CSV_MIME, JSON_MIME, PARQUET_MIME, XLSX_MIME})

_FILTER_OPERATORS = {
    "eq": "=",
    "ne": "<>",
    "gt": ">",
    "gte": ">=",
    "lt": "<",
    "lte": "<=",
}
_METRICS = frozenset({"count", "sum", "avg", "min", "max"})
_OUTPUT_MIME_TYPES = {
    "csv": CSV_MIME,
    "json": JSON_MIME,
    "xlsx": XLSX_MIME,
    "parquet": PARQUET_MIME,
}


def _identifier(name: str, available: set[str]) -> str:
    if name not in available:
        raise BuiltinToolError("UNKNOWN_COLUMN", f"표에 없는 열입니다: {name}")
    return '"' + name.replace('"', '""') + '"'


def _alias(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _write_xlsx_as_csv(data: bytes, path: Path) -> None:
    validate_ooxml_package(data)
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    except Exception as exc:  # noqa: BLE001 - parser 내부 경로를 숨긴다.
        raise BuiltinToolError("INVALID_XLSX", "XLSX 표를 읽지 못했습니다.") from exc
    try:
        if len(workbook.sheetnames) != 1:
            raise BuiltinToolError(
                "ONE_SHEET_REQUIRED", "표 가공은 시트가 하나인 XLSX부터 지원합니다."
            )
        with path.open("w", encoding="utf-8", newline="") as target:
            writer = csv.writer(target)
            for row in workbook.active.iter_rows(values_only=True):
                writer.writerow(list(row))
    finally:
        workbook.close()


def _source_path(*, data: bytes, mime_type: str, directory: Path, name: str) -> tuple[Path, str]:
    if not data:
        raise BuiltinToolError("EMPTY_FILE", "빈 표 파일은 가공할 수 없습니다.")
    if len(data) > MAX_FILE_BYTES:
        raise BuiltinToolError("FILE_TOO_LARGE", "표 파일은 20MB 이하여야 합니다.")
    if mime_type not in SUPPORTED_TABLE_MIME_TYPES:
        raise BuiltinToolError("UNSUPPORTED_FORMAT", "CSV, JSON, XLSX, Parquet만 지원합니다.")
    if mime_type == XLSX_MIME:
        path = directory / f"{name}.csv"
        _write_xlsx_as_csv(data, path)
        return path, CSV_MIME
    suffix = {CSV_MIME: ".csv", JSON_MIME: ".json", PARQUET_MIME: ".parquet"}[mime_type]
    path = directory / f"{name}{suffix}"
    path.write_bytes(data)
    return path, mime_type


def _create_source(connection: duckdb.DuckDBPyConnection, *, table: str, path: Path, mime: str) -> None:
    readers = {
        CSV_MIME: "read_csv_auto(?, header=true)",
        JSON_MIME: "read_json_auto(?)",
        PARQUET_MIME: "read_parquet(?)",
    }
    try:
        connection.execute(f'CREATE TABLE "{table}" AS SELECT * FROM {readers[mime]}', [str(path)])
    except duckdb.Error as exc:
        raise BuiltinToolError("INVALID_TABLE", "표 파일의 열과 값을 읽지 못했습니다.") from exc


def _columns(connection: duckdb.DuckDBPyConnection, table: str) -> list[str]:
    return [str(row[0]) for row in connection.execute(f'DESCRIBE "{table}"').fetchall()]


def _as_dict_list(value: Any, *, field: str) -> list[dict[str, Any]]:
    """모델이 배열 대신 객체 하나나 잘못된 모양을 보내도 배열로 맞춘다.

    LLM 은 `filters`·`sort`·`metrics` 를 `[{...}]` 대신 `{...}` 하나로 자주 보낸다.
    스키마에 array 라고 적어도 그렇다 — 여기서 관대하게 받아 준다. 그래도 모양이
    안 맞으면 `TOOL_EXECUTION_FAILED`(내부 예외)가 아니라 사용자용 오류를 낸다.
    """

    if value is None:
        return []
    if isinstance(value, dict):
        value = [value]
    if not isinstance(value, list) or any(not isinstance(item, dict) for item in value):
        raise BuiltinToolError("INVALID_ARGUMENT", f"{field} 는 객체 배열이어야 합니다.")
    return value


def _normalize_metrics(metrics: Any) -> list[dict[str, Any]]:
    """집계 지표를 표준 모양 `[{column, function, as}]` 로 맞춘다.

    받아 주는 변형:
    - `{"공수": {"function": "avg", "alias": "평균"}}`  → column 을 키에서 가져온다
    - `[{"column": "공수", "function": "avg", "alias": "평균"}]` → `alias` 를 `as` 로
    - `{"column": "공수", "function": "avg"}` → 배열로 감싼다
    """

    if metrics is None:
        return []
    if isinstance(metrics, dict) and metrics and all(
        isinstance(v, dict) for v in metrics.values()
    ):
        metrics = [{"column": key, **spec} for key, spec in metrics.items()]
    normalized: list[dict[str, Any]] = []
    for item in _as_dict_list(metrics, field="metrics"):
        entry = dict(item)
        if "as" not in entry and "alias" in entry:
            entry["as"] = entry.pop("alias")
        normalized.append(entry)
    return normalized


def _where(filters: list[dict[str, Any]], available: set[str]) -> tuple[str, list[Any]]:
    clauses = []
    values: list[Any] = []
    for item in filters:
        column = _identifier(str(item.get("column") or ""), available)
        operator = str(item.get("operator") or "")
        value = item.get("value")
        if operator in _FILTER_OPERATORS:
            clauses.append(f"{column} {_FILTER_OPERATORS[operator]} ?")
            values.append(value)
        elif operator == "contains":
            clauses.append(f"CAST({column} AS VARCHAR) LIKE ? ESCAPE '!'")
            escaped = str(value).replace("!", "!!").replace("%", "!%").replace("_", "!_")
            values.append(f"%{escaped}%")
        elif operator == "in" and isinstance(value, list) and value:
            clauses.append(f"{column} IN ({', '.join('?' for _ in value)})")
            values.extend(value)
        elif operator in {"is_null", "not_null"}:
            clauses.append(f"{column} IS {'NOT ' if operator == 'not_null' else ''}NULL")
        else:
            raise BuiltinToolError("INVALID_FILTER", "지원하지 않는 표 필터입니다.")
    return (" WHERE " + " AND ".join(clauses) if clauses else ""), values


def _sort(sort: list[dict[str, Any]], available: set[str]) -> str:
    parts = []
    for item in sort:
        direction = str(item.get("direction") or "asc").lower()
        if direction not in {"asc", "desc"}:
            raise BuiltinToolError("INVALID_SORT", "정렬 방향은 asc 또는 desc여야 합니다.")
        parts.append(f"{_identifier(str(item.get('column') or ''), available)} {direction.upper()}")
    return " ORDER BY " + ", ".join(parts) if parts else ""


def _result(
    connection: duckdb.DuckDBPyConnection,
    *,
    query: str,
    parameters: list[Any],
    limit: int,
) -> tuple[list[str], list[list[Any]], bool]:
    if not 1 <= limit <= 5000:
        raise BuiltinToolError("INVALID_LIMIT", "결과 행 제한은 1~5,000 사이여야 합니다.")
    cursor = connection.execute(f"SELECT * FROM ({query}) AS result LIMIT {limit + 1}", parameters)
    columns = [item[0] for item in cursor.description]
    raw = cursor.fetchall()
    return columns, [[json_value(value) for value in row] for row in raw[:limit]], len(raw) > limit


def _export(columns: list[str], rows: list[list[Any]], output_format: str) -> bytes:
    if output_format == "json":
        return json.dumps([dict(zip(columns, row)) for row in rows], ensure_ascii=False).encode()
    if output_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        writer.writerows(rows)
        return output.getvalue().encode("utf-8-sig")
    if output_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(columns)
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    raise BuiltinToolError("UNSUPPORTED_OUTPUT", "CSV, JSON, XLSX, Parquet로만 저장할 수 있습니다.")


def _full_result(
    connection: duckdb.DuckDBPyConnection, query: str, parameters: list[Any]
) -> tuple[list[str], list[list[Any]]]:
    cursor = connection.execute(
        f"SELECT * FROM ({query}) AS export_source LIMIT {MAX_TABLE_EXPORT_ROWS + 1}",
        parameters,
    )
    columns = [item[0] for item in cursor.description]
    raw = cursor.fetchall()
    if len(raw) > MAX_TABLE_EXPORT_ROWS:
        raise BuiltinToolError(
            "TOO_MANY_EXPORT_ROWS", "결과 파일은 100,000행 이하로 만들 수 있습니다."
        )
    return columns, [[json_value(value) for value in row] for row in raw]


def transform_table(
    *,
    data: bytes,
    mime_type: str,
    operation: str,
    columns: list[str] | None = None,
    filters: list[dict[str, Any]] | None = None,
    sort: list[dict[str, Any]] | None = None,
    group_by: list[str] | None = None,
    metrics: list[dict[str, str]] | None = None,
    join_data: bytes | None = None,
    join_mime_type: str | None = None,
    join_on: list[dict[str, str]] | None = None,
    limit: int = 100,
    output_format: str | None = None,
) -> dict[str, Any]:
    """SQL을 입력받지 않고 허용된 표 operation만 실행한다."""

    if output_format is not None and output_format not in _OUTPUT_MIME_TYPES:
        raise BuiltinToolError("UNSUPPORTED_OUTPUT", "CSV, JSON, XLSX, Parquet로만 저장할 수 있습니다.")

    # 모델이 배열 대신 객체 하나나 문자열을 보내는 흔한 실수를 여기서 흡수한다.
    filters = _as_dict_list(filters, field="filters")
    sort = _as_dict_list(sort, field="sort")
    join_on = _as_dict_list(join_on, field="join_on")
    metrics = _normalize_metrics(metrics)
    if isinstance(columns, str):
        columns = [columns]
    if isinstance(group_by, str):
        group_by = [group_by]

    with TemporaryDirectory(prefix="builtin-table-") as raw_directory:
        directory = Path(raw_directory)
        source_path, source_mime = _source_path(
            data=data, mime_type=mime_type, directory=directory, name="source"
        )
        connection = duckdb.connect(":memory:")
        try:
            _create_source(connection, table="source", path=source_path, mime=source_mime)
            available_columns = _columns(connection, "source")
            available = set(available_columns)
            selected = columns or available_columns
            select_sql = ", ".join(_identifier(name, available) for name in selected)
            where_sql, parameters = _where(filters or [], available)
            sort_sql = _sort(sort or [], available)

            if operation in {"preview", "filter", "sort", "convert"}:
                query = f'SELECT {select_sql} FROM "source"{where_sql}{sort_sql}'
            elif operation == "aggregate":
                groups = group_by or []
                group_sql = [_identifier(name, available) for name in groups]
                metric_sql = []
                for metric in metrics or []:
                    function = str(metric.get("function") or "").lower()
                    if function not in _METRICS:
                        raise BuiltinToolError("INVALID_METRIC", "지원하지 않는 집계 함수입니다.")
                    column_name = str(metric.get("column") or "")
                    column_sql = "*" if function == "count" and not column_name else _identifier(column_name, available)
                    alias = str(metric.get("as") or f"{function}_{column_name or 'rows'}")
                    metric_sql.append(f'{function.upper()}({column_sql}) AS "{alias.replace(chr(34), chr(34) * 2)}"')
                if not metric_sql:
                    raise BuiltinToolError("METRIC_REQUIRED", "집계할 값을 하나 이상 지정해 주세요.")
                query = f'SELECT {", ".join(group_sql + metric_sql)} FROM "source"{where_sql}'
                if group_sql:
                    query += " GROUP BY " + ", ".join(group_sql)
                query += sort_sql
            elif operation == "statistics":
                if sort:
                    raise BuiltinToolError(
                        "INVALID_STATISTICS_SORT", "기본 통계 결과에는 원본 열 정렬을 적용할 수 없습니다."
                    )
                query = f'SUMMARIZE SELECT {select_sql} FROM "source"{where_sql}'
            elif operation == "join":
                if join_data is None or join_mime_type is None or not join_on:
                    raise BuiltinToolError("JOIN_INPUT_REQUIRED", "결합할 두 번째 표와 기준 열이 필요합니다.")
                right_path, right_mime = _source_path(
                    data=join_data,
                    mime_type=join_mime_type,
                    directory=directory,
                    name="right_source",
                )
                _create_source(connection, table="right_source", path=right_path, mime=right_mime)
                right_columns = _columns(connection, "right_source")
                right_available = set(right_columns)
                conditions = [
                    f'l.{_identifier(str(item.get("left") or ""), available)} = '
                    f'r.{_identifier(str(item.get("right") or ""), right_available)}'
                    for item in join_on
                ]
                left_select = [
                    f"l.{_identifier(name, available)} AS {_alias(f'left.{name}')}"
                    for name in available_columns
                ]
                right_select = [
                    f"r.{_identifier(name, right_available)} AS {_alias(f'right.{name}')}"
                    for name in right_columns
                ]
                query = (
                    f'SELECT {", ".join(left_select + right_select)} FROM "source" l '
                    f'INNER JOIN "right_source" r ON {" AND ".join(conditions)}'
                )
                parameters = []
            else:
                raise BuiltinToolError("UNSUPPORTED_OPERATION", "지원하지 않는 표 가공 작업입니다.")

            result_columns, rows, truncated = _result(
                connection, query=query, parameters=parameters, limit=limit
            )
            result: dict[str, Any] = {
                "columns": result_columns,
                "rows": rows,
                "row_count": len(rows),
                "truncated": truncated,
            }
            if output_format:
                export_columns, export_rows = _full_result(connection, query, parameters)
                if output_format == "parquet":
                    connection.execute(f"CREATE TEMP TABLE export_result AS {query}", parameters)
                    target = directory / "result.parquet"
                    escaped_target = str(target).replace("'", "''")
                    connection.execute(
                        f"COPY export_result TO '{escaped_target}' (FORMAT PARQUET)"
                    )
                    output = target.read_bytes()
                else:
                    output = _export(export_columns, export_rows, output_format)
                if len(output) > MAX_OUTPUT_BYTES:
                    raise BuiltinToolError("OUTPUT_TOO_LARGE", "결과 파일이 허용 크기를 초과했습니다.")
                result["output_bytes"] = output
                result["output_mime_type"] = _OUTPUT_MIME_TYPES[output_format]
            return result
        except duckdb.Error as exc:
            raise BuiltinToolError("TABLE_OPERATION_FAILED", "표 가공을 완료하지 못했습니다.") from exc
        finally:
            connection.close()
